import numpy as np

# Initialization
S = 48      # Slots number
P = 8      # Pole number
Nph = 3     # Phase number

Check = np.zeros((S,P), dtype=bool)
for i in range(2,S):
    for j in range(2,P): 
        # Calculate K0
        K0 = np.zeros(int(i/2))
        for q in range(int(i/2)):
            K0[q] = 2*i/Nph/j*(1+Nph*q)

        # Check any int in K0
        Any_int = any(float(x).is_integer() for x in np.nditer(K0))
        
        Check[i-1,j-1] = Any_int
print('Check = ')
print(Check)

from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 將 NumPy 陣列轉換為 Python 列表
Check = Check.tolist()

# 創建綠色填充樣式
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# 創建一個新的 Excel 工作簿
workbook = Workbook()

# 獲取當前工作表
worksheet = workbook.active

# 將數據寫入工作表中
worksheet.append(range(0,P+1))  # 追加第一行
for i, row in enumerate(Check, start=1):
    row_with_index = [i] + row
    worksheet.append(row_with_index)

    # 設置底部顏色為綠色
    for col_num, value in enumerate(row_with_index, start=1):
        cell = worksheet.cell(row=worksheet.max_row, column=col_num)
        if col_num == len(row_with_index) and value:
            cell.fill = green_fill

# 將數據保存到文件中
workbook.save('Balanced-Motor-Matrix.xlsx')